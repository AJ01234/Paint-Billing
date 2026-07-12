from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, time
import json
import io
from pathlib import Path
import threading
import time as time_module
from typing import Any
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen
import base64

from fastapi import Depends, FastAPI, Form, HTTPException, Query, Request
from fastapi import UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.exc import OperationalError
from sqlmodel import Session, delete, or_, select

from .billing import InvoiceLineInput, build_product_search_text, calculate_invoice, matches_product_search
from .config import SHOP_DETAILS
from .db import create_db_and_tables, engine, get_session, seed_data
from .models import (
    BillType,
    DailyReport,
    ExpenseEntry,
    ExpenseType,
    Invoice,
    InvoiceItem,
    InvoiceStatus,
    LedgerEntry,
    LedgerEntryType,
    NotificationChannel,
    NotificationLog,
    NotificationStatus,
    OwnerSettings,
    Party,
    PaymentReminder,
    PartyType,
    Product,
    ReminderStatus,
)
from .settings import SETTINGS


BASE_DIR = Path(__file__).resolve().parent
REPORTS_DIR = BASE_DIR.parent / "generated_reports"
INVOICES_DIR = BASE_DIR.parent / "generated_invoices"
ACCOUNTING_DIR = BASE_DIR.parent / "generated_accounting"
app = FastAPI(title=SHOP_DETAILS.app_name)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def currency(value: float) -> str:
    return f"{value:,.2f}"


templates.env.filters["currency"] = currency
scheduler_started = False


@app.on_event("startup")
def on_startup() -> None:
    create_db_and_tables()
    seed_data()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    INVOICES_DIR.mkdir(parents=True, exist_ok=True)
    ACCOUNTING_DIR.mkdir(parents=True, exist_ok=True)
    start_scheduler_if_enabled()


@app.get("/healthz")
def healthz() -> dict[str, str]:
    return {"status": "ok"}


def context(request: Request, **extra: Any) -> dict[str, Any]:
    return {"request": request, "shop": SHOP_DETAILS, "today": date.today(), **extra}


def base_url(request: Request) -> str:
    return SETTINGS.public_base_url or str(request.base_url).rstrip("/")


def normalize_phone(phone: str) -> str:
    digits = "".join(ch for ch in (phone or "") if ch.isdigit())
    if len(digits) == 10:
        return f"91{digits}"
    return digits


def build_whatsapp_share_url(phone: str, message: str, file_url: str = "") -> str:
    normalized_phone = normalize_phone(phone)
    full_message = message.strip()
    if file_url:
        full_message = f"{full_message}\n{file_url}"
    if not normalized_phone:
        return ""
    return f"https://wa.me/{normalized_phone}?text={quote(full_message)}"


def can_send_sms() -> bool:
    return all(
        [
            SETTINGS.twilio_account_sid,
            SETTINGS.twilio_auth_token,
            SETTINGS.twilio_from_phone,
        ]
    )


def send_sms_via_twilio(to_phone: str, body: str) -> None:
    account_sid = SETTINGS.twilio_account_sid
    auth_token = SETTINGS.twilio_auth_token
    from_phone = SETTINGS.twilio_from_phone
    normalized_to = normalize_phone(to_phone)
    if not normalized_to:
        raise RuntimeError("Recipient phone missing for Twilio SMS send")
    credentials = base64.b64encode(f"{account_sid}:{auth_token}".encode("utf-8")).decode("utf-8")
    payload = urlencode(
        {
            "From": from_phone,
            "To": f"+{normalized_to}",
            "Body": body,
        }
    ).encode("utf-8")
    request = Request(
        f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json",
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Basic {credentials}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )
    with urlopen(request, timeout=20) as response:
        if response.status >= 400:
            raise RuntimeError(f"Twilio SMS send failed with status {response.status}")


def get_owner_settings(session: Session) -> OwnerSettings:
    try:
        settings = session.exec(select(OwnerSettings).order_by(OwnerSettings.id)).first()
    except OperationalError:
        create_db_and_tables()
        settings = session.exec(select(OwnerSettings).order_by(OwnerSettings.id)).first()
    if settings:
        return settings
    settings = OwnerSettings(
        owner_name="Anklikar Owner",
        primary_phone=SHOP_DETAILS.phone,
        whatsapp_phone=SHOP_DETAILS.phone,
        report_start_hour=7,
        report_end_hour=23,
        enable_sms=False,
        enable_whatsapp=False,
    )
    session.add(settings)
    session.commit()
    session.refresh(settings)
    return settings


def queue_notification(
    session: Session,
    *,
    channel: NotificationChannel,
    recipient_name: str,
    recipient_phone: str,
    message_type: str,
    message_body: str,
    file_url: str = "",
    invoice_id: int | None = None,
    report_id: int | None = None,
    status: NotificationStatus = NotificationStatus.pending,
) -> None:
    session.add(
        NotificationLog(
            channel=channel,
            status=status,
            recipient_name=recipient_name,
            recipient_phone=recipient_phone,
            message_type=message_type,
            message_body=message_body,
            file_url=file_url,
            invoice_id=invoice_id,
            report_id=report_id,
        )
    )


def process_pending_notifications(session: Session) -> None:
    settings = get_owner_settings(session)
    notifications = session.exec(
        select(NotificationLog).where(NotificationLog.status == NotificationStatus.pending).order_by(NotificationLog.created_at.asc())
    ).all()
    for notification in notifications:
        if notification.channel == NotificationChannel.sms:
            if not settings.enable_sms or not can_send_sms() or not notification.recipient_phone:
                continue
            try:
                send_sms_via_twilio(notification.recipient_phone, notification.message_body + (f" {notification.file_url}" if notification.file_url else ""))
                notification.status = NotificationStatus.sent
            except Exception:
                notification.status = NotificationStatus.failed
            session.add(notification)


def queue_owner_report_dispatch(session: Session, target_date: date, settings: OwnerSettings, public_base: str) -> None:
    report = generate_daily_report(session, target_date, settings)
    owner_phone = settings.primary_phone
    whatsapp_phone = settings.whatsapp_phone or settings.primary_phone
    report_url = f"{public_base}/reports/{report.id}/download" if public_base else f"/reports/{report.id}/download"
    already_queued = session.exec(
        select(NotificationLog).where(
            NotificationLog.report_id == report.id,
            NotificationLog.message_type == "DAILY_REPORT_DISPATCH",
        )
    ).first()
    if already_queued:
        return
    summary = report.report_summary
    if owner_phone and settings.enable_sms and can_send_sms():
        queue_notification(
            session,
            channel=NotificationChannel.sms,
            recipient_name=settings.owner_name,
            recipient_phone=owner_phone,
            message_type="DAILY_REPORT_DISPATCH",
            message_body=summary,
            file_url=report_url,
            report_id=report.id,
        )
    if whatsapp_phone:
        queue_notification(
            session,
            channel=NotificationChannel.whatsapp,
            recipient_name=settings.owner_name,
            recipient_phone=whatsapp_phone,
            message_type="DAILY_REPORT_DISPATCH",
            message_body=summary,
            file_url=report_url,
            report_id=report.id,
        )


def scheduler_loop() -> None:
    while True:
        try:
            with Session(engine) as session:
                settings = get_owner_settings(session)
                now = datetime.now()
                if now.hour >= settings.report_end_hour:
                    public_base = SETTINGS.public_base_url
                    queue_owner_report_dispatch(session, now.date(), settings, public_base)
                    process_pending_notifications(session)
                    session.commit()
        except Exception:
            pass
        time_module.sleep(max(30, SETTINGS.scheduler_poll_seconds))


def start_scheduler_if_enabled() -> None:
    global scheduler_started
    if scheduler_started or not SETTINGS.scheduler_enabled:
        return
    scheduler_started = True
    thread = threading.Thread(target=scheduler_loop, daemon=True, name="billing-scheduler")
    thread.start()


def build_invoice_share_summary(invoice: Invoice, party: Party | None) -> str:
    party_name = party.name if party else "Walk-in Customer"
    return "\n".join(
        [
            f"Namaste {party_name},",
            f"Your quotation from {SHOP_DETAILS.business_name} is ready.",
            f"Total: Rs {currency(invoice.grand_total)}",
            f"Payment mode: {invoice.payment_mode}",
            f"Status: {invoice.status.value}",
            "Please keep this PDF quotation for your records.",
        ]
    )


def build_report_share_summary(report: DailyReport) -> str:
    return "\n".join(
        [
            f"Daily report for {report.report_date}",
            f"Sales: Rs {currency(report.total_sales)}",
            f"Invoices: {report.invoice_count}",
            f"Outstanding ledger: Rs {currency(report.outstanding_ledger)}",
            f"Low stock items: {report.low_stock_count}",
        ]
    )


def report_window(target_date: date, settings: OwnerSettings) -> tuple[datetime, datetime]:
    return (
        datetime.combine(target_date, time(hour=settings.report_start_hour)),
        datetime.combine(target_date, time(hour=settings.report_end_hour)),
    )


def generate_daily_report(session: Session, target_date: date, settings: OwnerSettings) -> DailyReport:
    period_start, period_end = report_window(target_date, settings)
    invoices = session.exec(
        select(Invoice).where(
            Invoice.status == InvoiceStatus.finalized,
            Invoice.invoice_date >= period_start,
            Invoice.invoice_date <= period_end,
        )
    ).all()
    low_stock_count = len(
        session.exec(select(Product).where(Product.stock_quantity <= Product.low_stock_threshold)).all()
    )
    outstanding_ledger = round(
        sum(party.running_balance for party in session.exec(select(Party)).all() if party.running_balance > 0),
        2,
    )
    top_lines: dict[str, float] = defaultdict(float)
    if invoices:
        invoice_ids = [invoice.id for invoice in invoices if invoice.id is not None]
        item_groups = get_invoice_items(session, invoice_ids)
        for items in item_groups.values():
            for item in items:
                top_lines[item.product_name] += item.quantity
    top_products = sorted(top_lines.items(), key=lambda pair: pair[1], reverse=True)[:5]
    summary = (
        f"{target_date.isoformat()} report: "
        f"{len(invoices)} invoices, "
        f"Rs {currency(sum(invoice.grand_total for invoice in invoices))} total sales, "
        f"{low_stock_count} low stock items, "
        f"Rs {currency(outstanding_ledger)} outstanding ledger."
    )
    report_payload = {
        "report_date": target_date.isoformat(),
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "invoice_count": len(invoices),
        "total_sales": round(sum(invoice.grand_total for invoice in invoices), 2),
        "outstanding_ledger": outstanding_ledger,
        "low_stock_count": low_stock_count,
        "top_products": [{"name": name, "quantity": qty} for name, qty in top_products],
    }
    report_path = REPORTS_DIR / f"daily-report-{target_date.isoformat()}.pdf"
    build_daily_report_pdf(
        report_path,
        report_date=target_date,
        payload=report_payload,
        settings=settings,
    )

    report_key = target_date.isoformat()
    report = session.exec(select(DailyReport).where(DailyReport.report_date == report_key)).first()
    is_new = report is None
    if report is None:
        report = DailyReport(report_date=report_key, period_start=period_start, period_end=period_end)
    report.period_start = period_start
    report.period_end = period_end
    report.invoice_count = len(invoices)
    report.total_sales = round(sum(invoice.grand_total for invoice in invoices), 2)
    report.outstanding_ledger = outstanding_ledger
    report.low_stock_count = low_stock_count
    report.report_summary = summary
    report.report_path = str(report_path)
    session.add(report)
    session.flush()

    if is_new:
        owner_phone = settings.whatsapp_phone or settings.primary_phone
        if owner_phone:
            queue_notification(
                session,
                channel=NotificationChannel.internal,
                recipient_name=settings.owner_name,
                recipient_phone=owner_phone,
                message_type="DAILY_REPORT",
                message_body=summary,
                file_url=f"/reports/{report.id}/download",
                report_id=report.id,
            )
    return report


def ensure_reports(session: Session) -> list[DailyReport]:
    settings = get_owner_settings(session)
    dates_to_generate = {
        invoice.invoice_date.date()
        for invoice in session.exec(select(Invoice).where(Invoice.status == InvoiceStatus.finalized)).all()
    }
    if not dates_to_generate:
        dates_to_generate.add(date.today())
    reports = [generate_daily_report(session, target_date, settings) for target_date in sorted(dates_to_generate, reverse=True)]
    session.commit()
    return sorted(reports, key=lambda report: report.report_date, reverse=True)


def draw_label_value(pdf: canvas.Canvas, x: float, y: float, label: str, value: str, *, bold: bool = False) -> None:
    pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
    pdf.drawString(x, y, label)
    pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
    pdf.drawString(x + 90, y, value)


def wrap_pdf_text(pdf: canvas.Canvas, text: str, max_width: float, font_name: str, font_size: float) -> list[str]:
    text = (text or "").strip()
    if not text:
        return [""]
    words = text.split()
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if pdf.stringWidth(trial, font_name, font_size) <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def draw_pdf_wrapped_lines(
    pdf: canvas.Canvas,
    *,
    x: float,
    y: float,
    text: str,
    max_width: float,
    font_name: str,
    font_size: float,
    line_height: float,
) -> float:
    lines = wrap_pdf_text(pdf, text, max_width, font_name, font_size)
    pdf.setFont(font_name, font_size)
    current_y = y
    for line in lines:
        pdf.drawString(x, current_y, line)
        current_y -= line_height
    return current_y


def build_invoice_pdf(
    pdf_path: Path,
    *,
    invoice: Invoice,
    items: list[InvoiceItem],
    party: Party | None,
    item_units: list[str],
    total_quantity: float,
    amount_words: str,
    party_pan: str,
) -> None:
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    left = 12 * mm
    right = width - 12 * mm
    top = height - 12 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, top, f"GSTIN : {SHOP_DETAILS.gstin}")
    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawRightString(right, top, "Original Copy")

    y = top - 12
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawCentredString(width / 2, y, "QUOTATION")
    y -= 14
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawCentredString(width / 2, y, f"M/S {SHOP_DETAILS.business_name}")
    y -= 12
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(width / 2, y, SHOP_DETAILS.address)

    y -= 12
    pdf.rect(left, y - 56, right - left, 56)
    mid = left + (right - left) / 2
    pdf.line(mid, y, mid, y - 56)
    draw_label_value(pdf, left + 4, y - 12, "Dated", invoice.invoice_date.strftime("%d-%m-%Y"))
    draw_label_value(pdf, left + 4, y - 24, "Place of Supply", f"{SHOP_DETAILS.state} ({SHOP_DETAILS.state_code})")
    draw_label_value(pdf, left + 4, y - 36, "Reverse Charge", "N")
    draw_label_value(pdf, left + 4, y - 48, "GR/RR No.", "")
    draw_label_value(pdf, mid + 4, y - 12, "Transport", "")
    draw_label_value(pdf, mid + 4, y - 24, "Vehicle No.", "")
    draw_label_value(pdf, mid + 4, y - 36, "Station", "")
    draw_label_value(pdf, mid + 4, y - 48, "E-Way Bill No.", "")

    y -= 56
    party_box_height = 90
    pdf.rect(left, y - party_box_height, right - left, party_box_height)
    pdf.line(mid, y, mid, y - party_box_height)
    pdf.setFont("Helvetica-BoldOblique", 10)
    pdf.drawString(left + 4, y - 12, "Billed to :")
    pdf.drawString(mid + 4, y - 12, "Shipped to :")
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left + 4, y - 24, party.name if party else "Walk-in Customer")
    pdf.drawString(mid + 4, y - 24, party.name if party else "Walk-in Customer")
    address_width = (mid - left) - 12
    party_address = party.address if party and party.address else ""
    draw_pdf_wrapped_lines(
        pdf,
        x=left + 4,
        y=y - 36,
        text=party_address,
        max_width=address_width,
        font_name="Helvetica",
        font_size=8.5,
        line_height=10,
    )
    draw_pdf_wrapped_lines(
        pdf,
        x=mid + 4,
        y=y - 36,
        text=party_address,
        max_width=address_width,
        font_name="Helvetica",
        font_size=8.5,
        line_height=10,
    )
    pdf.setFont("Helvetica", 9)
    if party and party.phone:
        pdf.drawString(left + 4, y - 56, party.phone)
        pdf.drawString(mid + 4, y - 56, party.phone)
    pdf.drawString(left + 4, y - 68, f"Party PAN : {party_pan}")
    pdf.drawString(mid + 4, y - 68, f"Party PAN : {party_pan}")
    pdf.drawString(left + 4, y - 80, f"GSTIN / UIN : {party.gstin if party and party.gstin else ''}")
    pdf.drawString(mid + 4, y - 80, f"GSTIN / UIN : {party.gstin if party and party.gstin else ''}")

    y -= party_box_height
    table_top = y
    row_height = 16
    col_widths = [18 * mm, 64 * mm, 24 * mm, 18 * mm, 18 * mm, 28 * mm, 30 * mm]
    headers = ["S.N.", "Description of Goods", "HSN/SAC\nCode", "Qty.", "Unit", "Price", "Amount(`)"]
    x = left
    pdf.rect(left, table_top - row_height, sum(col_widths), row_height)
    for idx, header in enumerate(headers):
        pdf.line(x, table_top, x, table_top - row_height)
        pdf.setFont("Helvetica-Bold", 8)
        for offset, line in enumerate(header.split("\n")):
            pdf.drawCentredString(x + col_widths[idx] / 2, table_top - 10 - (offset * 8), line)
        x += col_widths[idx]
    pdf.line(x, table_top, x, table_top - row_height)

    current_y = table_top - row_height
    pdf.setFont("Helvetica", 8.5)
    for idx, item in enumerate(items, start=1):
        description = f"{(item.brand + ' ' + item.product_name).strip().upper()}"
        desc_lines = wrap_pdf_text(pdf, description, col_widths[1] - 8, "Helvetica", 8)
        current_row_height = max(row_height, 10 + (len(desc_lines) * 8))
        pdf.rect(left, current_y - current_row_height, sum(col_widths), current_row_height)
        row_values = [
            f"{idx}.",
            description,
            item.hsn_code,
            f"{item.quantity:.2f}",
            item_units[idx - 1],
            f"{item.unit_price:.2f}",
            currency(item.taxable_value),
        ]
        x = left
        for col_idx, value in enumerate(row_values):
            pdf.line(x, current_y, x, current_y - current_row_height)
            if col_idx in {0, 2, 3, 5, 6}:
                pdf.drawRightString(x + col_widths[col_idx] - 4, current_y - 11, value)
            elif col_idx == 1:
                text_y = current_y - 11
                pdf.setFont("Helvetica", 8)
                for line in desc_lines:
                    pdf.drawString(x + 4, text_y, line)
                    text_y -= 8
            else:
                pdf.drawString(x + 4, current_y - 11, value)
            x += col_widths[col_idx]
        pdf.line(x, current_y, x, current_y - current_row_height)
        current_y -= current_row_height

    pdf.rect(left, current_y - 14, sum(col_widths), 14)
    pdf.drawRightString(right - 4, current_y - 10, currency(invoice.taxable_total))
    current_y -= 14

    pdf.rect(left, current_y - 50, sum(col_widths), 50)
    amount_box_x = right - 42 * mm
    pdf.line(amount_box_x, current_y, amount_box_x, current_y - 50)
    tax_lines = []
    if invoice.cgst:
        tax_lines.append(("Add : CGST", f"{items[0].gst_percent / 2:.2f} %", currency(invoice.cgst)))
    if invoice.sgst:
        tax_lines.append(("Add : SGST", f"{items[0].gst_percent / 2:.2f} %", currency(invoice.sgst)))
    if invoice.igst:
        tax_lines.append(("Add : IGST", f"{items[0].gst_percent:.2f} %", currency(invoice.igst)))
    tax_lines.append(("Add : Rounded Off (+)", "", currency(invoice.round_off)))
    line_y = current_y - 12
    for label, pct, amount in tax_lines:
        pdf.drawRightString(amount_box_x - 60, line_y, label)
        if pct:
            pdf.drawString(amount_box_x - 45, line_y, "@")
            pdf.drawRightString(amount_box_x - 8, line_y, pct)
        pdf.drawRightString(right - 4, line_y, amount)
        line_y -= 12
    current_y -= 50

    pdf.rect(left, current_y - 16, sum(col_widths), 16)
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawRightString(left + 100 * mm, current_y - 11, "Grand Total")
    pdf.drawRightString(left + 122 * mm, current_y - 11, f"{total_quantity:.2f}")
    pdf.drawString(left + 124 * mm, current_y - 11, item_units[0] if item_units else "Units")
    pdf.drawCentredString(left + 146 * mm, current_y - 11, "`")
    pdf.drawRightString(right - 4, current_y - 11, currency(invoice.grand_total))
    current_y -= 16

    pdf.rect(left, current_y - 26, sum(col_widths), 26)
    tax_headers = ["Tax Rate", "Taxable Amt.", "CGST Amt.", "SGST Amt.", "Total Tax"]
    tax_values = [
        f"{items[0].gst_percent if items else 0}%",
        currency(invoice.taxable_total),
        currency(invoice.cgst),
        currency(invoice.sgst),
        currency(invoice.igst if invoice.igst else invoice.cgst + invoice.sgst),
    ]
    tx_width = (right - left) / 5
    for idx, value in enumerate(tax_headers):
        x = left + idx * tx_width
        pdf.drawCentredString(x + tx_width / 2, current_y - 9, value)
        pdf.drawCentredString(x + tx_width / 2, current_y - 20, tax_values[idx])
    current_y -= 26

    pdf.rect(left, current_y - 16, sum(col_widths), 16)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left + 4, current_y - 11, amount_words)
    current_y -= 16

    footer_height = 62
    pdf.rect(left, current_y - footer_height, sum(col_widths), footer_height)
    pdf.line(mid, current_y, mid, current_y - footer_height)
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(left + 4, current_y - 10, "Terms & Conditions")
    pdf.setFont("Helvetica", 8)
    footer_lines = [
        "E.& O.E.",
        "1. Goods once sold will not be taken back.",
        "2. Interest @ 18% p.a. will be charged if payment is delayed.",
        f"3. Subject to '{SHOP_DETAILS.state}' Jurisdiction only.",
    ]
    for idx, line in enumerate(footer_lines):
        pdf.drawString(left + 4, current_y - 22 - (idx * 10), line)
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(mid + 4, current_y - 10, "Receiver's Signature :")
    pdf.drawCentredString(mid + 95, current_y - 42, f"for M/S {SHOP_DETAILS.business_name}")
    pdf.drawRightString(right - 4, current_y - 56, "Authorised Signatory")
    pdf.showPage()
    pdf.save()


def build_daily_report_pdf(pdf_path: Path, *, report_date: date, payload: dict[str, Any], settings: OwnerSettings) -> None:
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    left = 16 * mm
    top = height - 18 * mm
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(left, top, "Daily Billing Report")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left, top - 16, f"Shop: {SHOP_DETAILS.business_name}")
    pdf.drawString(left, top - 30, f"Date: {report_date.isoformat()}")
    pdf.drawString(left, top - 44, f"Window: {settings.report_start_hour}:00 to {settings.report_end_hour}:00")

    y = top - 72
    summary_lines = [
        ("Invoice count", str(payload["invoice_count"])),
        ("Total sales", f"Rs {currency(payload['total_sales'])}"),
        ("Outstanding ledger", f"Rs {currency(payload['outstanding_ledger'])}"),
        ("Low stock items", str(payload["low_stock_count"])),
    ]
    for label, value in summary_lines:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left, y, label)
        pdf.setFont("Helvetica", 11)
        pdf.drawString(left + 110, y, value)
        y -= 18

    y -= 12
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Top products")
    y -= 16
    top_products = payload.get("top_products", [])
    if not top_products:
        pdf.setFont("Helvetica", 10)
        pdf.drawString(left, y, "No product sales recorded in this report window.")
    else:
        for idx, item in enumerate(top_products, start=1):
            pdf.setFont("Helvetica", 10)
            pdf.drawString(left, y, f"{idx}. {item['name']}")
            pdf.drawRightString(width - 20 * mm, y, str(item["quantity"]))
            y -= 14

    y -= 20
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left, y, "Report summary")
    y -= 16
    pdf.setFont("Helvetica", 10)
    for chunk in [
        f"Generated for {report_date.isoformat()}",
        f"Owner notifications target: {settings.primary_phone or 'Not set'}",
        f"Low stock alerts and payment reminders are tracked inside the app.",
    ]:
        pdf.drawString(left, y, chunk)
        y -= 14

    pdf.showPage()
    pdf.save()


ONES = [
    "Zero", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
    "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
    "Seventeen", "Eighteen", "Nineteen",
]
TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def number_to_words(value: int) -> str:
    if value < 20:
        return ONES[value]
    if value < 100:
        return TENS[value // 10] + (f" {ONES[value % 10]}" if value % 10 else "")
    if value < 1000:
        return ONES[value // 100] + " Hundred" + (f" {number_to_words(value % 100)}" if value % 100 else "")
    if value < 100000:
        return number_to_words(value // 1000) + " Thousand" + (f" {number_to_words(value % 1000)}" if value % 1000 else "")
    if value < 10000000:
        return number_to_words(value // 100000) + " Lakh" + (f" {number_to_words(value % 100000)}" if value % 100000 else "")
    return number_to_words(value // 10000000) + " Crore" + (f" {number_to_words(value % 10000000)}" if value % 10000000 else "")


def amount_in_words(amount: float) -> str:
    rounded = int(round(amount))
    return f"Rupees {number_to_words(rounded)} Only"


def extract_pan(gstin: str) -> str:
    cleaned = (gstin or "").strip()
    return cleaned[2:12] if len(cleaned) >= 12 else ""


def infer_unit_label(size: str, product_name: str) -> str:
    source = f"{size} {product_name}".upper()
    if "KG" in source or "KGS" in source:
        return "Kgs."
    if "LTR" in source or "L" in source:
        return "KLR"
    return "Pcs."


def split_tax(item: InvoiceItem, supply_type: str) -> dict[str, float]:
    if supply_type == "INTER_STATE":
        return {
            "cgst_rate": 0.0,
            "cgst_amount": 0.0,
            "sgst_rate": 0.0,
            "sgst_amount": 0.0,
            "igst_rate": item.gst_percent,
            "igst_amount": item.gst_amount,
        }
    return {
        "cgst_rate": round(item.gst_percent / 2.0, 2),
        "cgst_amount": round(item.gst_amount / 2.0, 2),
        "sgst_rate": round(item.gst_percent / 2.0, 2),
        "sgst_amount": round(item.gst_amount / 2.0, 2),
        "igst_rate": 0.0,
        "igst_amount": 0.0,
    }


def make_invoice_number(session: Session) -> str:
    today_code = datetime.now().strftime("%Y%m%d")
    prefix = f"INV-{today_code}-"
    existing = session.exec(select(Invoice).where(Invoice.invoice_number.startswith(prefix))).all()
    sequence = len(existing) + 1
    return f"{prefix}{sequence:03d}"


def get_party_map(session: Session) -> dict[int, Party]:
    return {party.id: party for party in session.exec(select(Party).order_by(Party.name)).all() if party.id is not None}


def format_product_display(item: InvoiceItem | Product) -> str:
    return " | ".join(
        part for part in [getattr(item, "brand", ""), getattr(item, "product_name", getattr(item, "name", "")), getattr(item, "size", ""), getattr(item, "shade", "")] if part
    )


def get_invoice_items(session: Session, invoice_ids: list[int]) -> dict[int, list[InvoiceItem]]:
    if not invoice_ids:
        return {}
    grouped: dict[int, list[InvoiceItem]] = defaultdict(list)
    items = session.exec(select(InvoiceItem).where(InvoiceItem.invoice_id.in_(invoice_ids))).all()
    for item in items:
        grouped[item.invoice_id].append(item)
    return grouped


def restore_invoice_effects(session: Session, invoice: Invoice) -> None:
    items = session.exec(select(InvoiceItem).where(InvoiceItem.invoice_id == invoice.id)).all()
    for item in items:
        if item.product_id:
            product = session.get(Product, item.product_id)
            if product:
                product.stock_quantity = round(product.stock_quantity + item.quantity, 2)
                session.add(product)

    if invoice.party_id and invoice.payment_mode.lower() == "credit":
        party = session.get(Party, invoice.party_id)
        if party:
            party.running_balance = round(max(party.running_balance - invoice.grand_total, 0.0), 2)
            session.add(party)


def delete_invoice_children(session: Session, invoice_id: int) -> None:
    session.exec(delete(LedgerEntry).where(LedgerEntry.invoice_id == invoice_id))
    session.exec(delete(InvoiceItem).where(InvoiceItem.invoice_id == invoice_id))


def financial_year_bounds(target: date) -> tuple[date, date]:
    start_year = target.year if target.month >= 4 else target.year - 1
    start = date(start_year, 4, 1)
    end = date(start_year + 1, 3, 31)
    return start, end


def financial_year_label(target: date) -> str:
    start, end = financial_year_bounds(target)
    return f"FY {start.year}-{str(end.year)[-2:]}"


def daterange_to_datetimes(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    return datetime.combine(start_date, time.min), datetime.combine(end_date, time.max)


def monthly_bucket_key(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def infer_likely_itr(turnover: float, personal_expenses: float) -> tuple[str, str]:
    if turnover <= 5000000:
        return (
            "Likely ITR-4 (if presumptive scheme applies)",
            "This looks like a small business turnover range where presumptive taxation may apply if the owner qualifies as a resident individual/HUF/eligible firm and chooses presumptive business income.",
        )
    return (
        "Review ITR-3 / detailed filing",
        "Turnover is above the simple presumptive helper range used here, so treat this as a detailed filing case and verify the correct return on the Income Tax portal before submission.",
    )


def build_accounting_summary(
    session: Session,
    *,
    start_date: date,
    end_date: date,
) -> dict[str, Any]:
    start_dt, end_dt = daterange_to_datetimes(start_date, end_date)
    invoices = session.exec(
        select(Invoice).where(
            Invoice.status == InvoiceStatus.finalized,
            Invoice.invoice_date >= start_dt,
            Invoice.invoice_date <= end_dt,
        )
    ).all()
    expenses = session.exec(
        select(ExpenseEntry).where(
            ExpenseEntry.entry_date >= start_dt,
            ExpenseEntry.entry_date <= end_dt,
        ).order_by(ExpenseEntry.entry_date.desc())
    ).all()
    low_stock_items = session.exec(
        select(Product).where(Product.stock_quantity <= Product.low_stock_threshold)
    ).all()
    outstanding_ledger = round(
        sum(party.running_balance for party in session.exec(select(Party)).all() if party.running_balance > 0),
        2,
    )

    turnover = round(sum(invoice.grand_total for invoice in invoices), 2)
    taxable_sales = round(sum(invoice.taxable_total for invoice in invoices), 2)
    gst_collected = round(sum(invoice.cgst + invoice.sgst + invoice.igst for invoice in invoices), 2)
    invoice_discount_total = round(sum(invoice.discount_total for invoice in invoices), 2)
    business_expenses = round(sum(exp.amount for exp in expenses if exp.expense_type == ExpenseType.business), 2)
    personal_expenses = round(sum(exp.amount for exp in expenses if exp.expense_type == ExpenseType.personal), 2)
    net_profit = round(turnover - business_expenses, 2)
    cash_turnover = round(
        sum(invoice.grand_total for invoice in invoices if invoice.payment_mode.strip().lower() == "cash"),
        2,
    )
    digital_turnover = round(turnover - cash_turnover, 2)
    presumptive_income = round((cash_turnover * 0.08) + (digital_turnover * 0.06), 2)
    likely_itr, itr_note = infer_likely_itr(turnover, personal_expenses)

    monthly_totals: dict[str, dict[str, float]] = defaultdict(lambda: {"sales": 0.0, "expenses": 0.0})
    for invoice in invoices:
        monthly_totals[monthly_bucket_key(invoice.invoice_date)]["sales"] += invoice.grand_total
    for expense in expenses:
        monthly_totals[monthly_bucket_key(expense.entry_date)]["expenses"] += expense.amount
    monthly_rows = []
    for key in sorted(monthly_totals.keys(), reverse=True):
        month_date = datetime.strptime(key, "%Y-%m")
        sales = round(monthly_totals[key]["sales"], 2)
        total_expense = round(monthly_totals[key]["expenses"], 2)
        monthly_rows.append(
            {
                "label": month_date.strftime("%b %Y"),
                "sales": sales,
                "expenses": total_expense,
                "profit": round(sales - total_expense, 2),
            }
        )

    expense_breakdown: dict[str, float] = defaultdict(float)
    for expense in expenses:
        expense_breakdown[expense.category or "Misc"] += expense.amount
    expense_rows = [
        {"category": category, "amount": round(amount, 2)}
        for category, amount in sorted(expense_breakdown.items(), key=lambda pair: pair[1], reverse=True)
    ]

    return {
        "period_start": start_date,
        "period_end": end_date,
        "financial_year": financial_year_label(start_date),
        "turnover": turnover,
        "taxable_sales": taxable_sales,
        "gst_collected": gst_collected,
        "invoice_count": len(invoices),
        "invoice_discount_total": invoice_discount_total,
        "business_expenses": business_expenses,
        "personal_expenses": personal_expenses,
        "net_profit": net_profit,
        "cash_turnover": cash_turnover,
        "digital_turnover": digital_turnover,
        "presumptive_income": presumptive_income,
        "outstanding_ledger": outstanding_ledger,
        "low_stock_count": len(low_stock_items),
        "monthly_rows": monthly_rows,
        "expense_rows": expense_rows,
        "expenses": expenses,
        "likely_itr": likely_itr,
        "itr_note": itr_note,
    }


def build_accounting_pdf(
    pdf_path: Path,
    *,
    summary: dict[str, Any],
    owner_settings: OwnerSettings,
) -> None:
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
    width, height = A4
    left = 16 * mm
    top = height - 18 * mm
    y = top

    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(left, y, "Business Accounting & ITR Helper")
    y -= 16
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left, y, f"Shop: {SHOP_DETAILS.business_name}")
    y -= 12
    pdf.drawString(left, y, f"Period: {summary['period_start'].isoformat()} to {summary['period_end'].isoformat()} ({summary['financial_year']})")
    y -= 12
    pdf.drawString(left, y, f"Owner: {owner_settings.owner_name}")
    y -= 22

    blocks = [
        ("Gross turnover", f"Rs {currency(summary['turnover'])}"),
        ("Taxable sales", f"Rs {currency(summary['taxable_sales'])}"),
        ("GST collected", f"Rs {currency(summary['gst_collected'])}"),
        ("Business expenses", f"Rs {currency(summary['business_expenses'])}"),
        ("Personal expenses", f"Rs {currency(summary['personal_expenses'])}"),
        ("Net profit", f"Rs {currency(summary['net_profit'])}"),
        ("Cash turnover", f"Rs {currency(summary['cash_turnover'])}"),
        ("Digital turnover", f"Rs {currency(summary['digital_turnover'])}"),
        ("Presumptive income helper", f"Rs {currency(summary['presumptive_income'])}"),
        ("Outstanding ledger", f"Rs {currency(summary['outstanding_ledger'])}"),
    ]
    for label, value in blocks:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left, y, label)
        pdf.setFont("Helvetica", 11)
        pdf.drawRightString(width - 16 * mm, y, value)
        y -= 16

    y -= 6
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Likely filing path")
    y -= 14
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, y, summary["likely_itr"])
    y -= 12
    pdf.setFont("Helvetica", 9)
    for line in [summary["itr_note"], "Use this as a preparation pack only. Verify the actual return form and final figures on the official Income Tax portal before filing."]:
        pdf.drawString(left, y, line[:130])
        y -= 12

    y -= 10
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Monthly snapshot")
    y -= 16
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, "Month")
    pdf.drawString(left + 60 * mm, y, "Sales")
    pdf.drawString(left + 100 * mm, y, "Expenses")
    pdf.drawString(left + 145 * mm, y, "Profit")
    y -= 10
    pdf.setFont("Helvetica", 9)
    for row in summary["monthly_rows"][:8]:
        pdf.drawString(left, y, row["label"])
        pdf.drawRightString(left + 95 * mm, y, currency(row["sales"]))
        pdf.drawRightString(left + 140 * mm, y, currency(row["expenses"]))
        pdf.drawRightString(width - 16 * mm, y, currency(row["profit"]))
        y -= 11
        if y < 30 * mm:
            pdf.showPage()
            y = top
            pdf.setFont("Helvetica", 9)

    pdf.showPage()
    pdf.save()


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    today_start = datetime.combine(date.today(), time.min)
    today_end = datetime.combine(date.today(), time.max)
    owner_settings = get_owner_settings(session)

    finalized_invoices = session.exec(
        select(Invoice).where(Invoice.status == InvoiceStatus.finalized)
    ).all()
    today_invoices = [
        invoice
        for invoice in finalized_invoices
        if today_start <= invoice.invoice_date <= today_end
    ]

    today_sales = round(sum(invoice.grand_total for invoice in today_invoices), 2)
    outstanding_ledger = round(
        sum(party.running_balance for party in session.exec(select(Party)).all() if party.running_balance > 0), 2
    )
    low_stock_items = session.exec(
        select(Product).where(Product.stock_quantity <= Product.low_stock_threshold).order_by(Product.stock_quantity)
    ).all()

    invoice_ids = [invoice.id for invoice in finalized_invoices if invoice.id is not None]
    item_groups = get_invoice_items(session, invoice_ids)
    top_sellers: dict[str, float] = defaultdict(float)
    for items in item_groups.values():
        for item in items:
            top_sellers[item.product_name] += item.quantity
    top_products = sorted(top_sellers.items(), key=lambda pair: pair[1], reverse=True)[:5]

    recent_invoices = sorted(finalized_invoices, key=lambda invoice: invoice.invoice_date, reverse=True)[:8]
    reports = ensure_reports(session)[:10]
    pending_notifications = session.exec(
        select(NotificationLog).where(NotificationLog.status == NotificationStatus.pending).order_by(NotificationLog.created_at.desc())
    ).all()
    reminders = session.exec(
        select(PaymentReminder).where(PaymentReminder.status == ReminderStatus.pending).order_by(PaymentReminder.due_date)
    ).all()
    report_share_urls = {
        report.id: build_whatsapp_share_url(
            owner_settings.whatsapp_phone or owner_settings.primary_phone,
            build_report_share_summary(report),
            f"{base_url(request)}/reports/{report.id}/download",
        )
        for report in reports
        if report.id is not None
    }

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        context(
            request,
            metrics={
                "today_sales": today_sales,
                "today_invoice_count": len(today_invoices),
                "outstanding_ledger": outstanding_ledger,
                "low_stock_count": len(low_stock_items),
            },
            low_stock_items=low_stock_items,
            top_products=top_products,
            recent_invoices=recent_invoices,
            parties=get_party_map(session),
            recent_reports=reports,
            pending_notifications=pending_notifications[:8],
            pending_notification_count=len(pending_notifications),
            upcoming_reminders=reminders[:8],
            due_reminder_count=len(reminders),
            owner_settings=owner_settings,
            report_share_urls=report_share_urls,
        ),
    )


@app.get("/accounting", response_class=HTMLResponse)
def accounting_dashboard(
    request: Request,
    fy_year: int | None = Query(default=None),
    session: Session = Depends(get_session),
) -> HTMLResponse:
    anchor_date = date(fy_year, 4, 1) if fy_year else date.today()
    start_date, end_date = financial_year_bounds(anchor_date)
    summary = build_accounting_summary(session, start_date=start_date, end_date=end_date)
    owner_settings = get_owner_settings(session)
    expense_rows = summary["expense_rows"]
    accounting_pdf_path = ACCOUNTING_DIR / f"itr-helper-{start_date.year}-{end_date.year}.pdf"
    build_accounting_pdf(accounting_pdf_path, summary=summary, owner_settings=owner_settings)
    previous_fy_start = start_date.year - 1
    next_fy_start = start_date.year + 1
    return templates.TemplateResponse(
        request,
        "accounting.html",
        context(
            request,
            summary=summary,
            recent_expenses=summary["expenses"][:10],
            expense_rows=expense_rows[:8],
            previous_fy_start=previous_fy_start,
            next_fy_start=next_fy_start,
            can_go_next=next_fy_start <= date.today().year,
        ),
    )


@app.post("/expenses")
def create_expense(
    expense_date: str = Form(...),
    amount: float = Form(...),
    category: str = Form("Misc"),
    expense_type: str = Form("BUSINESS"),
    payment_mode: str = Form("Cash"),
    vendor_name: str = Form(""),
    notes: str = Form(""),
    session: Session = Depends(get_session),
):
    try:
        entry_date = datetime.strptime(expense_date, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid expense date") from exc
    expense = ExpenseEntry(
        entry_date=entry_date,
        amount=round(amount, 2),
        category=(category or "Misc").strip() or "Misc",
        expense_type=ExpenseType.personal if expense_type == ExpenseType.personal.value else ExpenseType.business,
        payment_mode=(payment_mode or "Cash").strip() or "Cash",
        vendor_name=vendor_name.strip(),
        notes=notes.strip(),
    )
    session.add(expense)
    session.commit()
    fy_start, _ = financial_year_bounds(entry_date.date())
    return RedirectResponse(url=f"/accounting?fy_year={fy_start.year}", status_code=303)


@app.get("/accounting/export")
def download_accounting_pack(
    fy_year: int | None = Query(default=None),
    session: Session = Depends(get_session),
):
    anchor_date = date(fy_year, 4, 1) if fy_year else date.today()
    start_date, end_date = financial_year_bounds(anchor_date)
    owner_settings = get_owner_settings(session)
    summary = build_accounting_summary(session, start_date=start_date, end_date=end_date)
    pdf_path = ACCOUNTING_DIR / f"itr-helper-{start_date.year}-{end_date.year}.pdf"
    build_accounting_pdf(pdf_path, summary=summary, owner_settings=owner_settings)
    return FileResponse(pdf_path, media_type="application/pdf", filename=pdf_path.name)


@app.get("/products", response_class=HTMLResponse)
def list_products(
    request: Request,
    q: str = "",
    session: Session = Depends(get_session),
) -> HTMLResponse:
    products = session.exec(select(Product).order_by(Product.brand, Product.name, Product.size)).all()
    if q.strip():
        products = [
            product
            for product in products
            if matches_product_search(
                q, build_product_search_text(product.name, product.brand, product.size, product.shade, product.category)
            )
        ]
    return templates.TemplateResponse(request, "products.html", context(request, products=products, q=q, import_result=None))


@app.post("/products")
def create_product(
    name: str = Form(...),
    brand: str = Form(""),
    category: str = Form(""),
    size: str = Form(""),
    shade: str = Form(""),
    hsn_code: str = Form(""),
    gst_percent: float = Form(18.0),
    selling_price: float = Form(...),
    stock_quantity: float = Form(0.0),
    low_stock_threshold: float = Form(5.0),
    barcode: str = Form(""),
    session: Session = Depends(get_session),
):
    session.add(
        Product(
            name=name.strip(),
            brand=brand.strip(),
            category=category.strip(),
            size=size.strip(),
            shade=shade.strip(),
            hsn_code=hsn_code.strip(),
            gst_percent=gst_percent,
            selling_price=selling_price,
            stock_quantity=stock_quantity,
            low_stock_threshold=low_stock_threshold,
            barcode=barcode.strip(),
        )
    )
    session.commit()
    return RedirectResponse("/products", status_code=303)


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def pick_value(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


@app.post("/products/import")
async def import_products(
    request: Request,
    file: UploadFile,
    session: Session = Depends(get_session),
):
    content = await file.read()
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [normalize_header(cell) for cell in rows[0]]
    created = 0
    updated = 0

    for raw_row in rows[1:]:
        if not any(raw_row):
            continue
        row = {headers[idx]: raw_row[idx] for idx in range(min(len(headers), len(raw_row)))}
        name = str(pick_value(row, "name", "productname", "product", default="")).strip()
        if not name:
            continue
        brand = str(pick_value(row, "brand", default="")).strip()
        size = str(pick_value(row, "size", "packsize", default="")).strip()
        shade = str(pick_value(row, "shade", "color", default="")).strip()
        category = str(pick_value(row, "category", "type", default="")).strip()
        hsn_code = str(pick_value(row, "hsn", "hsncode", "hsn/saccode", default="")).strip()
        gst_percent = float(pick_value(row, "gst", "gstpercent", "gst%", default=18) or 18)
        selling_price = float(pick_value(row, "price", "sellingprice", "rate", default=0) or 0)
        stock_quantity = float(pick_value(row, "stock", "stockquantity", "qty", default=0) or 0)
        low_stock_threshold = float(pick_value(row, "threshold", "lowstockthreshold", default=5) or 5)
        barcode = str(pick_value(row, "barcode", "code", "productcode", default="")).strip()

        existing = session.exec(
            select(Product).where(
                Product.name == name,
                Product.brand == brand,
                Product.size == size,
                Product.shade == shade,
            )
        ).first()
        if existing:
            existing.category = category
            existing.hsn_code = hsn_code
            existing.gst_percent = gst_percent
            existing.selling_price = selling_price or existing.selling_price
            existing.stock_quantity = stock_quantity
            existing.low_stock_threshold = low_stock_threshold
            existing.barcode = barcode
            session.add(existing)
            updated += 1
        else:
            session.add(
                Product(
                    name=name,
                    brand=brand,
                    category=category,
                    size=size,
                    shade=shade,
                    hsn_code=hsn_code,
                    gst_percent=gst_percent,
                    selling_price=selling_price,
                    stock_quantity=stock_quantity,
                    low_stock_threshold=low_stock_threshold,
                    barcode=barcode,
                )
            )
            created += 1

    session.commit()
    products = session.exec(select(Product).order_by(Product.brand, Product.name, Product.size)).all()
    return templates.TemplateResponse(
        request,
        "products.html",
        context(
            request,
            products=products,
            q="",
            import_result={"created": created, "updated": updated, "filename": file.filename or "uploaded.xlsx"},
        ),
    )


@app.get("/parties", response_class=HTMLResponse)
def list_parties(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    parties = session.exec(select(Party).order_by(Party.party_type, Party.name)).all()
    ledger_entries = session.exec(select(LedgerEntry).order_by(LedgerEntry.entry_date.desc())).all()
    reminders = session.exec(select(PaymentReminder).order_by(PaymentReminder.due_date)).all()
    return templates.TemplateResponse(
        request,
        "parties.html",
        context(request, parties=parties, ledger_entries=ledger_entries, reminders=reminders, party_map=get_party_map(session)),
    )


@app.post("/parties")
def create_party(
    name: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    gstin: str = Form(""),
    party_type: PartyType = Form(PartyType.customer),
    session: Session = Depends(get_session),
):
    session.add(
        Party(
            name=name.strip(),
            phone=phone.strip(),
            address=address.strip(),
            gstin=gstin.strip(),
            party_type=party_type,
        )
    )
    session.commit()
    return RedirectResponse("/parties", status_code=303)


@app.post("/ledger/payment")
def record_payment(
    party_id: int = Form(...),
    amount: float = Form(...),
    description: str = Form("Payment received"),
    session: Session = Depends(get_session),
):
    party = session.get(Party, party_id)
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    party.running_balance = round(max(party.running_balance - amount, 0.0), 2)
    session.add(
        LedgerEntry(
            party_id=party.id,
            entry_type=LedgerEntryType.credit,
            amount=amount,
            description=description.strip() or "Payment received",
        )
    )
    session.add(party)
    session.commit()
    return RedirectResponse("/parties", status_code=303)


@app.post("/reminders")
def create_reminder(
    party_id: int = Form(...),
    due_date: str = Form(...),
    amount_due: float = Form(...),
    notes: str = Form(""),
    invoice_id: int | None = Form(default=None),
    session: Session = Depends(get_session),
):
    due_at = datetime.fromisoformat(f"{due_date}T09:00:00")
    reminder = PaymentReminder(
        party_id=party_id,
        invoice_id=invoice_id,
        due_date=due_at,
        amount_due=amount_due,
        notes=notes.strip(),
        status=ReminderStatus.pending,
    )
    session.add(reminder)
    party = session.get(Party, party_id)
    if party and party.phone:
        queue_notification(
            session,
            channel=NotificationChannel.internal,
            recipient_name=party.name,
            recipient_phone=party.phone,
            message_type="PAYMENT_REMINDER",
            message_body=(
                f"Reminder scheduled for {party.name} on {due_at.strftime('%d %b %Y')} "
                f"for Rs {currency(amount_due)}. {notes.strip()}".strip()
            ),
            invoice_id=invoice_id,
        )
        process_pending_notifications(session)
    session.commit()
    return RedirectResponse("/parties", status_code=303)


@app.post("/settings/owner")
def update_owner_settings(
    owner_name: str = Form(...),
    primary_phone: str = Form(""),
    whatsapp_phone: str = Form(""),
    report_start_hour: int = Form(7),
    report_end_hour: int = Form(23),
    enable_sms: bool = Form(False),
    enable_whatsapp: bool = Form(False),
    session: Session = Depends(get_session),
):
    settings = get_owner_settings(session)
    settings.owner_name = owner_name.strip()
    settings.primary_phone = primary_phone.strip()
    settings.whatsapp_phone = whatsapp_phone.strip()
    settings.report_start_hour = report_start_hour
    settings.report_end_hour = report_end_hour
    settings.enable_sms = enable_sms
    settings.enable_whatsapp = enable_whatsapp
    session.add(settings)
    process_pending_notifications(session)
    session.commit()
    return RedirectResponse("/", status_code=303)


@app.get("/api/products/search")
def product_search(
    q: str = Query(default=""),
    session: Session = Depends(get_session),
) -> list[dict[str, Any]]:
    products = session.exec(select(Product).order_by(Product.brand, Product.name)).all()
    filtered = [
        product
        for product in products
        if matches_product_search(
            q, build_product_search_text(product.name, product.brand, product.size, product.shade, product.category)
        )
    ][:10]
    return [
        {
            "id": product.id,
            "name": product.name,
            "brand": product.brand,
            "size": product.size,
            "shade": product.shade,
            "price": product.selling_price,
            "gst_percent": product.gst_percent,
            "stock_quantity": product.stock_quantity,
            "hsn_code": product.hsn_code,
            "display": " | ".join(
                part for part in [product.brand, product.name, product.size, product.shade] if part
            ),
        }
        for product in filtered
    ]


@app.get("/invoices/new", response_class=HTMLResponse)
def new_invoice(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    parties = session.exec(select(Party).order_by(Party.name)).all()
    return templates.TemplateResponse(
        request,
        "invoice_form.html",
        context(
            request,
            parties=parties,
            invoice_number=make_invoice_number(session),
            submit_label="Save and open printable bill",
            form_action="/invoices",
            invoice_lines=[None],
            selected_party_id="",
            selected_bill_type="GST",
            selected_supply_type="INTRA_STATE",
            selected_payment_mode="Cash",
            existing_notes="",
            existing_custom_party_name="",
            existing_custom_party_phone="",
            existing_custom_party_gstin="",
            existing_custom_party_address="",
            existing_custom_party_type="customer",
            payment_modes=["Cash", "UPI", "Card", "Credit"],
            supply_types=[
                ("INTRA_STATE", "Within State (CGST + SGST)"),
                ("INTER_STATE", "Outside State (IGST)"),
            ],
        ),
    )


@app.get("/invoices/{invoice_id}/edit", response_class=HTMLResponse)
def edit_invoice_form(invoice_id: int, request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    items = session.exec(select(InvoiceItem).where(InvoiceItem.invoice_id == invoice_id)).all()
    parties = session.exec(select(Party).order_by(Party.name)).all()
    invoice_lines = [
        {
            "product_id": item.product_id,
            "display": format_product_display(item),
            "quantity": item.quantity,
            "discount": item.discount,
            "price": item.unit_price,
            "gst_percent": item.gst_percent,
        }
        for item in items
    ] or [None]
    return templates.TemplateResponse(
        request,
        "invoice_form.html",
        context(
            request,
            parties=parties,
            invoice_number=invoice.invoice_number,
            submit_label="Update bill",
            form_action=f"/invoices/{invoice_id}/edit",
            invoice_lines=invoice_lines,
            selected_party_id=invoice.party_id or "",
            selected_bill_type=invoice.bill_type.value,
            selected_supply_type=invoice.supply_type,
            selected_payment_mode=invoice.payment_mode,
            existing_notes=invoice.notes,
            existing_custom_party_name="",
            existing_custom_party_phone="",
            existing_custom_party_gstin="",
            existing_custom_party_address="",
            existing_custom_party_type="customer",
            payment_modes=["Cash", "UPI", "Card", "Credit"],
            supply_types=[
                ("INTRA_STATE", "Within State (CGST + SGST)"),
                ("INTER_STATE", "Outside State (IGST)"),
            ],
        ),
    )


@app.post("/invoices")
async def create_invoice(request: Request, session: Session = Depends(get_session)):
    form = await request.form()
    raw_party_id = str(form.get("party_id", "")).strip()
    party_id = int(raw_party_id) if raw_party_id.isdigit() else None
    bill_type = BillType(form.get("bill_type", BillType.gst.value))
    supply_type = str(form.get("supply_type", "INTRA_STATE")).strip() or "INTRA_STATE"
    payment_mode = str(form.get("payment_mode", "Cash")).strip()
    notes = str(form.get("notes", "")).strip()
    custom_party_name = str(form.get("custom_party_name", "")).strip()
    custom_party_phone = str(form.get("custom_party_phone", "")).strip()
    custom_party_address = str(form.get("custom_party_address", "")).strip()
    custom_party_gstin = str(form.get("custom_party_gstin", "")).strip()
    custom_party_type = PartyType(str(form.get("custom_party_type", PartyType.customer.value)).strip())

    raw_product_ids = form.getlist("product_id")
    quantities = form.getlist("quantity")
    discounts = form.getlist("discount")
    unit_prices = form.getlist("unit_price")

    party = session.get(Party, party_id) if party_id else None
    if custom_party_name:
        party = Party(
            name=custom_party_name,
            phone=custom_party_phone,
            address=custom_party_address,
            gstin=custom_party_gstin,
            party_type=custom_party_type,
        )
        session.add(party)
        session.flush()

    product_map = {
        product.id: product for product in session.exec(select(Product)).all() if product.id is not None
    }
    input_lines: list[InvoiceLineInput] = []
    selected_products: list[Product] = []

    for index, raw_product_id in enumerate(raw_product_ids):
        if not str(raw_product_id).strip():
            continue
        product_id = int(raw_product_id)
        product = product_map.get(product_id)
        if not product:
            raise HTTPException(status_code=400, detail=f"Invalid product at row {index + 1}")
        quantity = float(quantities[index] or 0)
        discount = float(discounts[index] or 0)
        unit_price = float(unit_prices[index] or product.selling_price)
        if quantity <= 0:
            raise HTTPException(status_code=400, detail=f"Quantity must be positive at row {index + 1}")
        if unit_price <= 0:
            raise HTTPException(status_code=400, detail=f"Price must be positive at row {index + 1}")
        if product.stock_quantity < quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock for {product.name}")
        input_lines.append(
            InvoiceLineInput(
                product_id=product.id,
                name=product.name,
                quantity=quantity,
                unit_price=unit_price,
                gst_percent=product.gst_percent,
                discount=discount,
            )
        )
        selected_products.append(product)

    if not input_lines:
        raise HTTPException(status_code=400, detail="Select at least one product before saving the invoice")

    calculated_lines, summary = calculate_invoice(input_lines, bill_type.value, supply_type=supply_type)
    invoice = Invoice(
        invoice_number=make_invoice_number(session),
        invoice_date=datetime.now(),
        bill_type=bill_type,
        payment_mode=payment_mode,
        party_id=party.id if party else None,
        notes=notes,
        supply_type=supply_type,
        subtotal=summary.subtotal,
        discount_total=summary.discount_total,
        taxable_total=summary.taxable_total,
        cgst=summary.cgst if bill_type == BillType.gst else 0.0,
        sgst=summary.sgst if bill_type == BillType.gst else 0.0,
        igst=summary.igst if bill_type == BillType.gst else 0.0,
        round_off=summary.round_off,
        grand_total=summary.grand_total,
        status=InvoiceStatus.finalized,
    )
    session.add(invoice)
    session.flush()

    for calculated_line, product in zip(calculated_lines, selected_products):
        product.stock_quantity = round(product.stock_quantity - calculated_line.quantity, 2)
        session.add(product)
        session.add(
            InvoiceItem(
                invoice_id=invoice.id,
                product_id=product.id,
                product_name=product.name,
                brand=product.brand,
                size=product.size,
                shade=product.shade,
                hsn_code=product.hsn_code,
                quantity=calculated_line.quantity,
                unit_price=calculated_line.unit_price,
                gst_percent=calculated_line.gst_percent,
                discount=calculated_line.discount,
                taxable_value=calculated_line.taxable_value,
                gst_amount=calculated_line.gst_amount,
                line_total=calculated_line.line_total,
            )
        )

    if party and payment_mode.lower() == "credit":
        party.running_balance = round(party.running_balance + summary.grand_total, 2)
        session.add(party)
        session.add(
            LedgerEntry(
                party_id=party.id,
                invoice_id=invoice.id,
                entry_type=LedgerEntryType.debit,
                amount=summary.grand_total,
                description=f"Credit sale against invoice {invoice.invoice_number}",
            )
        )

    if party and party.phone:
        invoice_url = f"{base_url(request)}/invoices/{invoice.id}/pdf"
        share_summary = build_invoice_share_summary(invoice, party)
        owner_settings = get_owner_settings(session)
        if owner_settings.enable_sms and can_send_sms():
            queue_notification(
                session,
                channel=NotificationChannel.sms,
                recipient_name=party.name,
                recipient_phone=party.phone,
                message_type="INVOICE_CREATED",
                message_body=share_summary,
                file_url=invoice_url,
                invoice_id=invoice.id,
            )
        queue_notification(
            session,
            channel=NotificationChannel.whatsapp,
            recipient_name=party.name,
            recipient_phone=party.phone,
            message_type="INVOICE_CREATED",
            message_body=share_summary,
            file_url=invoice_url,
            invoice_id=invoice.id,
        )
        process_pending_notifications(session)

    session.commit()
    return RedirectResponse(f"/invoices/{invoice.id}", status_code=303)


@app.post("/invoices/{invoice_id}/edit")
async def update_invoice(invoice_id: int, request: Request, session: Session = Depends(get_session)):
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    restore_invoice_effects(session, invoice)
    delete_invoice_children(session, invoice_id)
    session.flush()

    form = await request.form()
    raw_party_id = str(form.get("party_id", "")).strip()
    party_id = int(raw_party_id) if raw_party_id.isdigit() else None
    bill_type = BillType(form.get("bill_type", BillType.gst.value))
    supply_type = str(form.get("supply_type", "INTRA_STATE")).strip() or "INTRA_STATE"
    payment_mode = str(form.get("payment_mode", "Cash")).strip()
    notes = str(form.get("notes", "")).strip()
    custom_party_name = str(form.get("custom_party_name", "")).strip()
    custom_party_phone = str(form.get("custom_party_phone", "")).strip()
    custom_party_address = str(form.get("custom_party_address", "")).strip()
    custom_party_gstin = str(form.get("custom_party_gstin", "")).strip()
    custom_party_type = PartyType(str(form.get("custom_party_type", PartyType.customer.value)).strip())

    raw_product_ids = form.getlist("product_id")
    quantities = form.getlist("quantity")
    discounts = form.getlist("discount")
    unit_prices = form.getlist("unit_price")

    party = session.get(Party, party_id) if party_id else None
    if custom_party_name:
        party = Party(
            name=custom_party_name,
            phone=custom_party_phone,
            address=custom_party_address,
            gstin=custom_party_gstin,
            party_type=custom_party_type,
        )
        session.add(party)
        session.flush()

    product_map = {
        product.id: product for product in session.exec(select(Product)).all() if product.id is not None
    }
    input_lines: list[InvoiceLineInput] = []
    selected_products: list[Product] = []

    for index, raw_product_id in enumerate(raw_product_ids):
        if not str(raw_product_id).strip():
            continue
        product_id = int(raw_product_id)
        product = product_map.get(product_id)
        if not product:
            raise HTTPException(status_code=400, detail=f"Invalid product at row {index + 1}")
        quantity = float(quantities[index] or 0)
        discount = float(discounts[index] or 0)
        unit_price = float(unit_prices[index] or product.selling_price)
        if quantity <= 0:
            raise HTTPException(status_code=400, detail=f"Quantity must be positive at row {index + 1}")
        if unit_price <= 0:
            raise HTTPException(status_code=400, detail=f"Price must be positive at row {index + 1}")
        if product.stock_quantity < quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock for {product.name}")
        input_lines.append(
            InvoiceLineInput(
                product_id=product.id,
                name=product.name,
                quantity=quantity,
                unit_price=unit_price,
                gst_percent=product.gst_percent,
                discount=discount,
            )
        )
        selected_products.append(product)

    if not input_lines:
        raise HTTPException(status_code=400, detail="Select at least one product before updating the invoice")

    calculated_lines, summary = calculate_invoice(input_lines, bill_type.value, supply_type=supply_type)
    invoice.bill_type = bill_type
    invoice.payment_mode = payment_mode
    invoice.party_id = party.id if party else None
    invoice.notes = notes
    invoice.supply_type = supply_type
    invoice.subtotal = summary.subtotal
    invoice.discount_total = summary.discount_total
    invoice.taxable_total = summary.taxable_total
    invoice.cgst = summary.cgst if bill_type == BillType.gst else 0.0
    invoice.sgst = summary.sgst if bill_type == BillType.gst else 0.0
    invoice.igst = summary.igst if bill_type == BillType.gst else 0.0
    invoice.round_off = summary.round_off
    invoice.grand_total = summary.grand_total
    invoice.status = InvoiceStatus.finalized
    session.add(invoice)

    for calculated_line, product in zip(calculated_lines, selected_products):
        product.stock_quantity = round(product.stock_quantity - calculated_line.quantity, 2)
        session.add(product)
        session.add(
            InvoiceItem(
                invoice_id=invoice.id,
                product_id=product.id,
                product_name=product.name,
                brand=product.brand,
                size=product.size,
                shade=product.shade,
                hsn_code=product.hsn_code,
                quantity=calculated_line.quantity,
                unit_price=calculated_line.unit_price,
                gst_percent=calculated_line.gst_percent,
                discount=calculated_line.discount,
                taxable_value=calculated_line.taxable_value,
                gst_amount=calculated_line.gst_amount,
                line_total=calculated_line.line_total,
            )
        )

    if party and payment_mode.lower() == "credit":
        party.running_balance = round(party.running_balance + summary.grand_total, 2)
        session.add(party)
        session.add(
            LedgerEntry(
                party_id=party.id,
                invoice_id=invoice.id,
                entry_type=LedgerEntryType.debit,
                amount=summary.grand_total,
                description=f"Credit sale against invoice {invoice.invoice_number}",
            )
        )

    if party and party.phone:
        invoice_url = f"{base_url(request)}/invoices/{invoice.id}/pdf"
        share_summary = build_invoice_share_summary(invoice, party)
        owner_settings = get_owner_settings(session)
        if owner_settings.enable_sms and can_send_sms():
            queue_notification(
                session,
                channel=NotificationChannel.sms,
                recipient_name=party.name,
                recipient_phone=party.phone,
                message_type="INVOICE_UPDATED",
                message_body=share_summary,
                file_url=invoice_url,
                invoice_id=invoice.id,
            )
        queue_notification(
            session,
            channel=NotificationChannel.whatsapp,
            recipient_name=party.name,
            recipient_phone=party.phone,
            message_type="INVOICE_UPDATED",
            message_body=share_summary,
            file_url=invoice_url,
            invoice_id=invoice.id,
        )
        process_pending_notifications(session)

    session.commit()
    return RedirectResponse(f"/invoices/{invoice.id}", status_code=303)


@app.get("/invoices", response_class=HTMLResponse)
def list_invoices(
    request: Request,
    q: str = "",
    payment_mode: str = "",
    date_from: str = "",
    date_to: str = "",
    session: Session = Depends(get_session),
) -> HTMLResponse:
    query = select(Invoice).order_by(Invoice.invoice_date.desc())

    if payment_mode.strip():
        query = query.where(Invoice.payment_mode == payment_mode.strip())
    if date_from.strip():
        query = query.where(Invoice.invoice_date >= datetime.fromisoformat(f"{date_from}T00:00:00"))
    if date_to.strip():
        query = query.where(Invoice.invoice_date <= datetime.fromisoformat(f"{date_to}T23:59:59"))
    if q.strip():
        like_term = f"%{q.strip()}%"
        matched_party_ids = [
            party.id
            for party in session.exec(select(Party).where(Party.name.ilike(like_term))).all()
            if party.id is not None
        ]
        if matched_party_ids:
            query = query.where(or_(Invoice.invoice_number.ilike(like_term), Invoice.party_id.in_(matched_party_ids)))
        else:
            query = query.where(Invoice.invoice_number.ilike(like_term))

    invoices = session.exec(query).all()
    return templates.TemplateResponse(
        request,
        "invoices.html",
        context(
            request,
            invoices=invoices,
            parties=get_party_map(session),
            filters={"q": q, "payment_mode": payment_mode, "date_from": date_from, "date_to": date_to},
        ),
    )


@app.get("/reports", response_class=HTMLResponse)
def list_reports(request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    reports = ensure_reports(session)
    notifications = session.exec(select(NotificationLog).order_by(NotificationLog.created_at.desc())).all()
    settings = get_owner_settings(session)
    report_share_urls = {
        report.id: build_whatsapp_share_url(
            settings.whatsapp_phone or settings.primary_phone,
            build_report_share_summary(report),
            f"{base_url(request)}/reports/{report.id}/download",
        )
        for report in reports
        if report.id is not None
    }
    return templates.TemplateResponse(
        request,
        "reports.html",
        context(request, reports=reports, notifications=notifications[:20], report_share_urls=report_share_urls),
    )


@app.get("/reports/{report_id}/download")
def download_report(report_id: int, session: Session = Depends(get_session)):
    report = session.get(DailyReport, report_id)
    if not report or not report.report_path:
        raise HTTPException(status_code=404, detail="Report not found")
    report_file = Path(report.report_path)
    if not report_file.exists():
        settings = get_owner_settings(session)
        report = generate_daily_report(session, datetime.fromisoformat(report.report_date).date(), settings)
        session.commit()
        report_file = Path(report.report_path)
    return FileResponse(report_file, media_type="application/pdf", filename=report_file.name)


@app.get("/invoices/{invoice_id}", response_class=HTMLResponse)
def invoice_detail(invoice_id: int, request: Request, session: Session = Depends(get_session)) -> HTMLResponse:
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    items = session.exec(select(InvoiceItem).where(InvoiceItem.invoice_id == invoice_id)).all()
    party = session.get(Party, invoice.party_id) if invoice.party_id else None
    line_tax_breakdown = [split_tax(item, invoice.supply_type) for item in items]
    total_quantity = round(sum(item.quantity for item in items), 2)
    item_units = [infer_unit_label(item.size, item.product_name) for item in items]
    share_summary = build_invoice_share_summary(invoice, party)
    whatsapp_share_url = ""
    if party and party.phone:
        whatsapp_share_url = build_whatsapp_share_url(
            party.phone,
            share_summary,
            f"{base_url(request)}/invoices/{invoice.id}/pdf",
        )
    return templates.TemplateResponse(
        request,
        "invoice_print.html",
        context(
            request,
            invoice=invoice,
            items=items,
            party=party,
            line_tax_breakdown=line_tax_breakdown,
            item_units=item_units,
            total_quantity=total_quantity,
            round_off=invoice.round_off,
            amount_in_words=amount_in_words(invoice.grand_total),
            party_pan=extract_pan(party.gstin if party else ""),
            whatsapp_share_url=whatsapp_share_url,
            whatsapp_share_text=share_summary,
        ),
    )


@app.get("/invoices/{invoice_id}/pdf")
def invoice_pdf(invoice_id: int, session: Session = Depends(get_session)):
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    items = session.exec(select(InvoiceItem).where(InvoiceItem.invoice_id == invoice_id)).all()
    party = session.get(Party, invoice.party_id) if invoice.party_id else None
    total_quantity = round(sum(item.quantity for item in items), 2)
    item_units = [infer_unit_label(item.size, item.product_name) for item in items]
    safe_party = "".join(ch for ch in (party.name if party else "customer").lower().replace(" ", "-") if ch.isalnum() or ch == "-").strip("-") or "customer"
    pdf_filename = f"quotation-{invoice.invoice_date.strftime('%Y%m%d')}-{safe_party}.pdf"
    pdf_path = INVOICES_DIR / pdf_filename
    build_invoice_pdf(
        pdf_path,
        invoice=invoice,
        items=items,
        party=party,
        item_units=item_units,
        total_quantity=total_quantity,
        amount_words=amount_in_words(invoice.grand_total),
        party_pan=extract_pan(party.gstin if party else ""),
    )
    return FileResponse(pdf_path, media_type="application/pdf", filename=pdf_filename)


@app.post("/invoices/{invoice_id}/cancel")
def cancel_invoice(invoice_id: int, session: Session = Depends(get_session)):
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.status == InvoiceStatus.cancelled:
        return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)

    restore_invoice_effects(session, invoice)
    delete_invoice_children(session, invoice.id)

    if invoice.party_id and invoice.payment_mode.lower() == "credit":
        party = session.get(Party, invoice.party_id)
        if party:
            session.add(
                LedgerEntry(
                    party_id=party.id,
                    invoice_id=invoice.id,
                    entry_type=LedgerEntryType.credit,
                    amount=invoice.grand_total,
                    description=f"Cancellation of invoice {invoice.invoice_number}",
                )
            )

    invoice.status = InvoiceStatus.cancelled
    session.add(invoice)
    session.commit()
    return RedirectResponse(f"/invoices/{invoice_id}", status_code=303)


@app.post("/invoices/{invoice_id}/delete")
def delete_invoice(invoice_id: int, session: Session = Depends(get_session)):
    invoice = session.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    restore_invoice_effects(session, invoice)
    delete_invoice_children(session, invoice.id)
    session.delete(invoice)
    session.commit()
    return RedirectResponse("/invoices", status_code=303)
